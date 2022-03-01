import traceback
import os
import http
from flask import Flask, request
from werkzeug.wrappers import Response
import cv2
import numpy as np
from PIL import Image
from matplotlib import pyplot as plt
from colour import Color
import logging
from telegram import Update, ForceReply
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, CallbackContext
from openpyxl import Workbook, load_workbook
import os
import io
import re


app = Flask(__name__)
# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)

logger = logging.getLogger(__name__)


class BigImage:
    def __init__(self, path):
        img = Image.open(path)
        img.load()
        self.image = np.asarray(img, dtype="int32")
        self.left_edge = 0
        self.up_edge = 0
        self.right_edge = self.image.shape[1]
        self.down_edge = self.image.shape[0]

    def show_full_image(self, show=False, image=None):
        if image is None:
            image = self.image
        if show:
            plt.imshow(image, interpolation='nearest')
            plt.show()


def draw_arrow_and_rectangle(big_image, target_coordinates):
    xs = abs(int(target_coordinates["xs"] - big_image.right_edge / 3))
    ys = abs(int(target_coordinates["ys"] - big_image.right_edge / 3))
    xe = target_coordinates["xs"]
    ye = target_coordinates["ys"]
    tmp = Color("red").rgb
    rgb_color = (tmp[0] * 255, tmp[1] * 255, tmp[2] * 255)
    # Draw arrowed line, from xs,ys to xe,ye in black with thickness 20 pixels
    cv2.arrowedLine(big_image.image, (xs, ys), (xe, ye), rgb_color, 12)
    return cv2.rectangle(big_image.image, (target_coordinates["xs"], target_coordinates["ys"]), (target_coordinates["xe"], target_coordinates["ye"]), rgb_color, 12)


def load_worksheet(workbook_path, headers=None, overwrite=False):
    workbook_name = workbook_path.split("/")[-1]
    workbook_folder = workbook_path.replace(workbook_name, "")
    if workbook_name in os.listdir(workbook_folder) and not overwrite:
        workbook = load_workbook(workbook_path)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = workbook_name.replace(".xlsx", "")
        worksheet.append(headers)
        workbook.save(workbook_path)
    return workbook


def get_sheet_by_columns_names(sheet, headers_row=0):
    return {
        COL[headers_row].value.lower() if "item image " in COL[headers_row].value.lower() else COL[
            headers_row].value: Current
        for Current, COL in enumerate(
            sheet.iter_cols(1, sheet.max_column)
        ) if COL[headers_row].value
    }


def start(update: Update, context: CallbackContext) -> None:
    user = update.effective_user
    update.message.reply_markdown_v2(
        fr'Hi {user.mention_markdown_v2()}\!',
        reply_markup=ForceReply(selective=True),
    )


def help_command(update: Update, context: CallbackContext) -> None:
    update.message.reply_text('USE: /ratmap FractionName.png')


def get_art_fractions_and_locations(folder_name):
    global arts
    wb = load_worksheet(f"./{folder_name}/Result.xlsx")
    worksheet = wb.active
    columns_names = get_sheet_by_columns_names(worksheet)
    fractions_and_locations = {
        worksheet[i][columns_names["ImageName"]].value: {
            "xs": worksheet[i][columns_names["Location XS"]].value,
            "xe": worksheet[i][columns_names["Location XE"]].value,
            "ys": worksheet[i][columns_names["Location YS"]].value,
            "ye": worksheet[i][columns_names["Location YE"]].value,
        }
        for i, row in enumerate(worksheet.iter_rows(min_row=2), 2)
        if worksheet[i][columns_names["ImageName"]].value is not None
    }

    arts[folder_name] = fractions_and_locations
    wb.close()


def locate_fraction(update: Update, context: CallbackContext) -> None:
    try:
        print("Locating Fraction...")
        pattern = re.compile("/ratmap (.*)-(.*).png$")
        if not re.search(pattern, update.message.text):
            update.message.reply_text("Invalid command! Your command should follow the following pattern:\n/ratmap Art_name-fraction_name.png\n"
                                      "(Eg: /ratmap I_Fought_The_Law-img1.png)")
            return
        big_image_name = context.args[0].split("-")[0]
        if big_image_name not in arts.keys():
            if big_image_name in os.listdir(os.getcwd()):
                get_art_fractions_and_locations(big_image_name)
            else:
                update.message.reply_text("Invalid fraction name or the art has not been fed yet to me, report to @jalil_bm if you are sure about the fraction name")
                return
        if context.args[0] not in list(arts[big_image_name].keys()):
            update.message.reply_text("Invalid fraction name or the art has not been fed yet to me, report to @jalil_bm if you are sure about the fraction name")
            return
        big_image = [filename for filename in os.listdir('.') if filename.startswith(f"{big_image_name}") and os.path.isfile(f"./{filename}")][0]
        na = draw_arrow_and_rectangle(BigImage(f"./{big_image}"), arts[big_image_name][context.args[0]])
        PIL_image = Image.fromarray(np.uint8(na)).convert('RGB')
        img_byte_arr = io.BytesIO()
        PIL_image.save(img_byte_arr, format='JPEG', dpi=[30, 30])
        # PIL_image.save('result.jpeg', format='JPEG', dpi=[300, 300])
        img_byte_arr = img_byte_arr.getvalue()
        print("Sending Location...")
        update.message.reply_photo(photo=img_byte_arr, filename=big_image_name + "-LOCATION-" + context.args[0].split("-")[1], timeout=1000)
    except Exception:
        print(traceback.format_exc())
        update.message.reply_text("Something went wrong with me :'( Please contact @jalil_bm")


def unknown_message(update: Update, context: CallbackContext) -> None:
    CommandHandler("ratmap_help", help_command)


if __name__ == "__main__":
    arts = {}
    updater = Updater("5250985047:AAFjg_weLVhIKp0sDBEZ5GYGEpQuYVife_A")
    dispatcher = updater.dispatcher
    dispatcher.add_handler(CommandHandler("start", start))
    dispatcher.add_handler(CommandHandler("ratmap_help", help_command))
    dispatcher.add_handler(CommandHandler("ratmap", locate_fraction))
    dispatcher.add_handler(MessageHandler(Filters.text & ~Filters.command, unknown_message))
    @app.post("/")
    def index() -> Response:
        dispatcher.process_update(
            Update.de_json(request.get_json(force=True), bot))

        return "", http.HTTPStatus.NO_CONTENT

# https://console.cloud.google.com/compute/instances?project=telegrambot-342723