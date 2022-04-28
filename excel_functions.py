from openpyxl import Workbook, load_workbook
import os

def load_worksheet(workbook_path, headers=None, overwrite=False):
    workbook_name = workbook_path.split("/")[-1]
    workbook_folder = workbook_path.replace(workbook_name, "")
    if workbook_name in os.listdir(workbook_folder) and not overwrite:
        workbook = load_workbook(workbook_path)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = workbook_name.replace(".xlsx", "")
        worksheet.append(headers)
        workbook.save(workbook_path)
    return workbook


def get_sheet_by_columns_names(sheet, headers_row=0):
    return {
        COL[headers_row].value.lower() if "item image " in COL[headers_row].value.lower() else COL[
            headers_row].value: Current
        for Current, COL in enumerate(
            sheet.iter_cols(1, sheet.max_column)
        ) if COL[headers_row].value
    }
