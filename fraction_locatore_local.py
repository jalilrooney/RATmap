import traceback
from typing import Tuple, Optional
import cv2
import numpy as np
from PIL import Image
from matplotlib import pyplot as plt
from colour import Color
import logging
from telegram import Update, ForceReply, Chat, ChatMember, ParseMode, ChatMemberUpdated
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, CallbackContext, ChatMemberHandler
from openpyxl import Workbook, load_workbook
import os
import io
import re

# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)

logger = logging.getLogger(__name__)


class BigImage:
    def __init__(self, path):
        img = Image.open(path)
        img.load()
        self.image = np.asarray(img, dtype="int32")
        self.left_edge = 0
        self.up_edge = 0
        self.right_edge = self.image.shape[1]
        self.down_edge = self.image.shape[0]

    def show_full_image(self, show=False, image=None):
        if image is None:
            image = self.image
        if show:
            plt.imshow(image, interpolation='nearest')
            plt.show()


def draw_arrow_and_rectangle(big_image, target_coordinates):
    xs = abs(int(target_coordinates["xs"] - big_image.right_edge / 3))
    ys = abs(int(target_coordinates["ys"] - big_image.right_edge / 3))
    xe = target_coordinates["xs"]
    ye = target_coordinates["ys"]
    tmp = Color("red").rgb
    rgb_color = (tmp[0] * 255, tmp[1] * 255, tmp[2] * 255)
    # Draw arrowed line, from xs,ys to xe,ye in black with thickness 20 pixels
    cv2.arrowedLine(big_image.image, (xs, ys), (xe, ye), rgb_color, 12)
    return cv2.rectangle(big_image.image, (target_coordinates["xs"], target_coordinates["ys"]), (target_coordinates["xe"], target_coordinates["ye"]), rgb_color, 12)


def load_worksheet(workbook_path, headers=None, overwrite=False):
    workbook_name = workbook_path.split("/")[-1]
    workbook_folder = workbook_path.replace(workbook_name, "")
    if workbook_name in os.listdir(workbook_folder) and not overwrite:
        workbook = load_workbook(workbook_path)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = workbook_name.replace(".xlsx", "")
        worksheet.append(headers)
        workbook.save(workbook_path)
    return workbook


def get_sheet_by_columns_names(sheet, headers_row=0):
    return {
        COL[headers_row].value.lower() if "item image " in COL[headers_row].value.lower() else COL[
            headers_row].value: Current
        for Current, COL in enumerate(
            sheet.iter_cols(1, sheet.max_column)
        ) if COL[headers_row].value
    }


def check_message_source(update):
    return update["message"]['chat']['username'] in ('TelethonAccounts', 'RATbits', 'jalil_bm', 'RatGermany')
    # return True


def help_command(update: Update, context: CallbackContext) -> None:
    if not check_message_source(update):
        update.message.reply_text('Please message me from https://t.me/RATbits')
        return
    update.message.reply_text('USE: /ratmap FractionName.png')


def get_art_fractions_and_locations(folder_name):
    global arts
    wb = load_worksheet(f"./{folder_name}/Result.xlsx")
    worksheet = wb.active
    columns_names = get_sheet_by_columns_names(worksheet)
    fractions_and_locations = {
        worksheet[i][columns_names["ImageName"]].value: {
            "xs": worksheet[i][columns_names["Location XS"]].value,
            "xe": worksheet[i][columns_names["Location XE"]].value,
            "ys": worksheet[i][columns_names["Location YS"]].value,
            "ye": worksheet[i][columns_names["Location YE"]].value,
        }
        for i, row in enumerate(worksheet.iter_rows(min_row=2), 2)
        if worksheet[i][columns_names["ImageName"]].value is not None
    }

    arts[folder_name] = fractions_and_locations
    wb.close()


def locate_fraction(update: Update, context: CallbackContext) -> None:
    if not check_message_source(update):
        update.message.reply_text('Please message me from https://t.me/RATbits')
        return
    try:
        print("Locating Fraction...")
        pattern = re.compile("/ratmap (.*)-(.*).png$")
        if not re.search(pattern, update.message.text):
            update.message.reply_text("Invalid command! Your command should follow the following pattern:\n/ratmap Art_name-fraction_name.png\n"
                                      "(Eg: /ratmap I_Fought_The_Law-img1.png)")
            return
        big_image_name = context.args[0].split("-")[0]
        if big_image_name not in arts.keys():
            if big_image_name in os.listdir(os.getcwd()):
                get_art_fractions_and_locations(big_image_name)
            else:
                update.message.reply_text("Invalid fraction name or the art has not been fed yet to me, report to @jalil_bm if you are sure about the fraction name")
                return
        if context.args[0] not in list(arts[big_image_name].keys()):
            update.message.reply_text("Invalid fraction name or the art has not been fed yet to me, report to @jalil_bm if you are sure about the fraction name")
            return
        big_image = [filename for filename in os.listdir('.') if filename.startswith(f"{big_image_name}") and os.path.isfile(f"./{filename}")][0]
        na = draw_arrow_and_rectangle(BigImage(f"./{big_image}"), arts[big_image_name][context.args[0]])
        PIL_image = Image.fromarray(np.uint8(na)).convert('RGB')
        img_byte_arr = io.BytesIO()
        PIL_image.save(img_byte_arr, format='JPEG', dpi=[30, 30])
        # PIL_image.save('result.jpeg', format='JPEG', dpi=[300, 300])
        img_byte_arr = img_byte_arr.getvalue()
        print("Sending Location...")
        update.message.reply_photo(photo=img_byte_arr, filename=big_image_name + "-LOCATION-" + context.args[0].split("-")[1], timeout=1000)
    except Exception:
        print(traceback.format_exc())
        update.message.reply_text("Something went wrong with me :'( Please contact @jalil_bm")


def unknown_message(update: Update, context: CallbackContext) -> None:
    CommandHandler("ratmap_help", help_command)


def track_chats(update: Update, context: CallbackContext) -> None:
    """Tracks the chats the bot is in."""
    result = extract_status_change(update.my_chat_member)
    if result is None:
        return
    was_member, is_member = result

    # Let's check who is responsible for the change
    cause_name = update.effective_user.full_name

    # Handle chat types differently:
    chat = update.effective_chat
    if chat.type == Chat.PRIVATE:
        if not was_member and is_member:
            logger.info("%s started the bot", cause_name)
            context.bot_data.setdefault("user_ids", set()).add(chat.id)
        elif was_member and not is_member:
            logger.info("%s blocked the bot", cause_name)
            context.bot_data.setdefault("user_ids", set()).discard(chat.id)
    elif chat.type in [Chat.GROUP, Chat.SUPERGROUP]:
        if not was_member and is_member:
            logger.info("%s added the bot to the group %s", cause_name, chat.title)
            context.bot_data.setdefault("group_ids", set()).add(chat.id)
        elif was_member and not is_member:
            logger.info("%s removed the bot from the group %s", cause_name, chat.title)
            context.bot_data.setdefault("group_ids", set()).discard(chat.id)
    else:
        if not was_member and is_member:
            logger.info("%s added the bot to the channel %s", cause_name, chat.title)
            context.bot_data.setdefault("channel_ids", set()).add(chat.id)
        elif was_member and not is_member:
            logger.info("%s removed the bot from the channel %s", cause_name, chat.title)
            context.bot_data.setdefault("channel_ids", set()).discard(chat.id)


def extract_status_change(chat_member_update: ChatMemberUpdated,) -> Optional[Tuple[bool, bool]]:
    """Takes a ChatMemberUpdated instance and extracts whether the 'old_chat_member' was a member
    of the chat and whether the 'new_chat_member' is a member of the chat. Returns None, if
    the status didn't change.
    """
    status_change = chat_member_update.difference().get("status")
    old_is_member, new_is_member = chat_member_update.difference().get("is_member", (None, None))

    if status_change is None:
        return None

    old_status, new_status = status_change
    was_member = (
        old_status
        in [
            ChatMember.MEMBER,
            ChatMember.CREATOR,
            ChatMember.ADMINISTRATOR,
        ]
        or (old_status == ChatMember.RESTRICTED and old_is_member is True)
    )
    is_member = (
        new_status
        in [
            ChatMember.MEMBER,
            ChatMember.CREATOR,
            ChatMember.ADMINISTRATOR,
        ]
        or (new_status == ChatMember.RESTRICTED and new_is_member is True)
    )

    return was_member, is_member


def greet_chat_members(update: Update, context: CallbackContext) -> None:
    """Greets new users in chats and announces when someone leaves"""
    result = extract_status_change(update.chat_member)
    if result is None:
        return
    was_member, is_member = result
    cause_name = update.chat_member.from_user.mention_html()
    member_name = update.chat_member.new_chat_member.user.mention_html()
    if not was_member and is_member:
        context.bot.send_message(
            chat_id=1055241434,
            text=f"{member_name} has joined the group...",
            parse_mode=ParseMode.HTML)
    elif was_member and not is_member:
        context.bot.send_message(
            chat_id=1055241434,
            text=f"{member_name} is no longer with us. Thanks a lot, {cause_name}...",
            parse_mode=ParseMode.HTML)


def show_chats(update: Update, context: CallbackContext) -> None:
    """Shows which chats the bot is in"""
    user_ids = ", ".join(str(uid) for uid in context.bot_data.setdefault("user_ids", set()))
    group_ids = ", ".join(str(gid) for gid in context.bot_data.setdefault("group_ids", set()))
    channel_ids = ", ".join(str(cid) for cid in context.bot_data.setdefault("channel_ids", set()))
    text = (
        f"@{context.bot.username} is currently in a conversation with the user IDs {user_ids}."
        f" Moreover it is a member of the groups with IDs {group_ids} "
        f"and administrator in the channels with IDs {channel_ids}."
    )
    print(text)
    update.effective_message.reply_text(text)



if __name__ == "__main__":
    arts = {}
    updater = Updater("5250985047:AAFjg_weLVhIKp0sDBEZ5GYGEpQuYVife_A")
    dispatcher = updater.dispatcher
    dispatcher.add_handler(CommandHandler("ratmap_help", help_command))
    dispatcher.add_handler(CommandHandler("ratmap", locate_fraction))
    dispatcher.add_handler(CommandHandler("show_chats", show_chats))
    dispatcher.add_handler(MessageHandler(Filters.text & ~Filters.command, unknown_message))
    dispatcher.add_handler(MessageHandler(Filters.text, unknown_message))
    dispatcher.add_handler(ChatMemberHandler(track_chats, ChatMemberHandler.MY_CHAT_MEMBER))
    dispatcher.add_handler(ChatMemberHandler(greet_chat_members, ChatMemberHandler.CHAT_MEMBER))
    updater.start_polling(allowed_updates=Update.ALL_TYPES)
    updater.start_polling()
    updater.idle()

# https://console.cloud.google.com/compute/instances?project=telegrambot-342723
# https://programmingforgood.medium.com/deploy-telegram-bot-on-google-cloud-platform-74f1f531f65e
# */1 * * * * cd /home/benharkatdjalil/RATmap; nohup python3 /home/benharkatdjalil/RATmap/fraction_locatore_local.py </dev/null &>/dev/null &
# chat id of djolocatorbo: 1055241434